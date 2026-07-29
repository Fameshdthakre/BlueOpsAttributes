import io
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from backend.database import get_connection
import json

_GREY = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
_BOLD = Font(bold=True)

def generate_excel_from_session(session_id: str) -> bytes:
    conn = get_connection()
    try:
        query = """
            SELECT 
                asin, attribute_id, product_type, brand, title,
                final_value, match_status, provider_used, confidence,
                raw_ai_value, extra_data, validated_product_type, validated_allowed_options
            FROM job_results
            WHERE session_id = %s
            ORDER BY id ASC
        """
        df = pd.read_sql(query, conn, params=(session_id,))
    finally:
        conn.close()

    if df.empty:
        raise ValueError(f"No results found for session {session_id}")

    rows = []
    for _, row_data in df.iterrows():
        # Parse extra_data back into dict if it exists
        extra = row_data["extra_data"]
        extra_dict = {}
        if extra:
            if isinstance(extra, dict):
                extra_dict = extra
            elif isinstance(extra, str) and extra.startswith("{"):
                try:
                    extra_dict = json.loads(extra)
                except Exception:
                    pass

        row = {
            "ASIN": row_data["asin"],
            "Attribute ID": row_data["attribute_id"],
            "Product Type": row_data["product_type"] or "",
            "Brand": row_data["brand"] or "",
            "Title": row_data["title"] or "",
            "Barcode": extra_dict.get("barcode", ""),
            "Description": extra_dict.get("description", ""),
            "Ref. Product Type": row_data["validated_product_type"] or "",
            "Ref. Allowed Options": row_data["validated_allowed_options"] or "",
            "Final Value": row_data["final_value"],
            "Match Status": row_data["match_status"],
            "Provider Used": row_data["provider_used"],
            "Confidence": f"{row_data['confidence']:.0%}" if pd.notnull(row_data['confidence']) else "0%",
            "Raw AI Response": row_data["raw_ai_value"],
        }
        
        # Add any remaining extra columns
        for k, v in extra_dict.items():
            if k not in row and k not in ("barcode", "description"):
                row[k] = v
                
        rows.append(row)

    out_df = pd.DataFrame(rows)
    
    # Save to BytesIO
    output = io.BytesIO()
    out_df.to_excel(output, index=False, engine="openpyxl")
    output.seek(0)
    
    # Apply formatting
    wb = load_workbook(output)
    ws = wb.active

    # Bold header with light grey background
    for cell in ws[1]:
        cell.font = _BOLD
        cell.fill = _GREY

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    final_output = io.BytesIO()
    wb.save(final_output)
    return final_output.getvalue()
