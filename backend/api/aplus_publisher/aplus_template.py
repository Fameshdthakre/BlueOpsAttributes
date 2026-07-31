import io
from fastapi import APIRouter, Header, HTTPException
from fastapi.responses import Response
from pydantic import BaseModel
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from backend.auth import verify_token
from backend.aplus_modules import MODULE_REGISTRY

router = APIRouter()

class TemplateRequest(BaseModel):
    module_ids: List[str]

def get_module(mod_id):
    for mod in MODULE_REGISTRY:
        if mod["id"] == mod_id:
            return mod
    return None

def apply_style(cell, font=None, fill=None, alignment=None, border=None):
    if font: cell.font = font
    if fill: cell.fill = fill
    if alignment: cell.alignment = alignment
    if border: cell.border = border
    else:
        thin = Side(border_style="thin", color="D9D9D9")
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

@router.post("/api/aplus/template")
def generate_aplus_template(req: TemplateRequest, x_user_id: int = Header(None), x_blueops_token: str = Header(None)):
    user_id = verify_token(x_user_id, x_blueops_token)
    if not user_id:
        raise HTTPException(status_code=401, detail="Unauthorized")

    if not req.module_ids:
        req.module_ids = ["module-5"]

    modules = [get_module(mid) for mid in req.module_ids if get_module(mid)]
    if not modules:
        raise HTTPException(status_code=400, detail="Invalid module IDs")

    wb = Workbook()
    
    # Styles
    f_doc_title = Font(bold=True, size=14, color="1F497D")
    f_doc_sub = Font(italic=True, size=10, color="555555")
    f_tbl_hdr = Font(bold=True, color="FFFFFF", size=11)
    fill_tbl_hdr = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    f_side_lbl = Font(bold=True, color="2C3E50", size=10)
    fill_side_lbl = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
    fill_config = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fill_highlight = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    no_border = Border(top=Side(style=None), bottom=Side(style=None), left=Side(style=None), right=Side(style=None))

    # Instructions Sheet
    ws_instr = wb.active
    ws_instr.title = "Instructions"
    ws_instr.column_dimensions['A'].width = 30
    ws_instr.column_dimensions['B'].width = 80
    
    instr_data = [
        ("A+ Publisher Pro - Template Instructions", "", f_doc_title, None),
        ("Each sheet tab represents one A+ module. Fill in the values and import the file back into the extension.", "", f_doc_sub, None),
        ("", "", None, None),
        ("Standard Fields (Row 7-8)", "Content Title and Draft URL are global for a module block. Keep them empty to auto-generate a Draft.", f_side_lbl, fill_side_lbl),
        ("Toggles (TRUE/FALSE)", "For settings like 'Show Reviews', type TRUE or FALSE in the designated cells.", f_side_lbl, fill_side_lbl),
        ("Repetitive Fields", "If a module has blocks (e.g. 4 images), fill them from left to right as Columns.", f_side_lbl, fill_side_lbl),
        ("Multiple Instances", "To publish 5 charts of the same module type, simply copy-paste the entire module block downward on the same sheet. Leave 2-3 empty rows between blocks.", f_side_lbl, fill_side_lbl),
    ]

    for row_idx, row_info in enumerate(instr_data, 1):
        c1 = ws_instr.cell(row=row_idx, column=1, value=row_info[0])
        c2 = ws_instr.cell(row=row_idx, column=2, value=row_info[1])
        if row_idx <= 3:
            apply_style(c1, font=row_info[2], border=no_border)
            apply_style(c2, border=no_border)
        else:
            apply_style(c1, font=row_info[2], fill=row_info[3], alignment=align_left)
            apply_style(c2, font=Font(size=10, color="333333"), alignment=align_left)

    # Sheets Map to group same modules on one sheet
    sheets_map = {}
    for mod in modules:
        if mod["id"] not in sheets_map:
            sheets_map[mod["id"]] = {"mod": mod, "count": 0}
        sheets_map[mod["id"]]["count"] += 1

    for mod_id, info in sheets_map.items():
        mod = info["mod"]
        ws = wb.create_sheet(title=mod["shortName"][:31])
        
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col_letter].width = 25 if col_letter == 'A' else 20

        is_comp = (mod_id == "module-5")
        current_row = 1
        
        for instance in range(info["count"]):
            if instance > 0:
                current_row += 3 # spacing
            
            c = ws.cell(row=current_row, column=1, value=f"{mod['name']} Data"); apply_style(c, font=f_doc_title, border=no_border); current_row += 1
            c = ws.cell(row=current_row, column=1, value=f"Module Type: {mod['id']}"); apply_style(c, font=f_doc_sub, border=no_border); current_row += 1
            current_row += 1 # Empty row
            
            # Header
            c = ws.cell(row=current_row, column=1, value="Field Name"); apply_style(c, font=f_tbl_hdr, fill=fill_tbl_hdr, alignment=align_center)
            c = ws.cell(row=current_row, column=2, value="Value"); apply_style(c, font=f_tbl_hdr, fill=fill_tbl_hdr, alignment=align_center)
            current_row += 1
            
            c = ws.cell(row=current_row, column=1, value="Content Title"); apply_style(c, font=f_side_lbl, fill=fill_side_lbl, alignment=align_left)
            c = ws.cell(row=current_row, column=2, value=""); apply_style(c, alignment=align_center)
            current_row += 1
            
            c = ws.cell(row=current_row, column=1, value="Draft URL"); apply_style(c, font=f_side_lbl, fill=fill_side_lbl, alignment=align_left)
            c = ws.cell(row=current_row, column=2, value=""); apply_style(c, alignment=align_center)
            current_row += 1
            
            if is_comp:
                current_row += 1
                headers = ["Configuration", "Base Product", "Competitor 1", "Competitor 2", "Competitor 3", "Competitor 4", "Competitor 5"]
                for i, h in enumerate(headers, 1):
                    c = ws.cell(row=current_row, column=i, value=h); apply_style(c, font=f_tbl_hdr, fill=fill_tbl_hdr, alignment=align_center)
                current_row += 1
                
                # Comparison specific stuff
                c = ws.cell(row=current_row, column=1, value="Highlight Column"); apply_style(c, font=f_side_lbl, fill=fill_side_lbl, alignment=align_left)
                for i in range(2, 8):
                    c = ws.cell(row=current_row, column=i, value=False); apply_style(c, fill=fill_config, alignment=align_center)
                current_row += 1
                
                for toggle in ["Show Reviews", "Show Prices", "Show Add To Cart"]:
                    c = ws.cell(row=current_row, column=1, value=toggle); apply_style(c, font=f_side_lbl, fill=fill_side_lbl, alignment=align_left)
                    c = ws.cell(row=current_row, column=2, value=True); apply_style(c, fill=fill_config, alignment=align_center)
                    current_row += 1
                    
                current_row += 1
                headers = ["Products", "Base Product", "Competitor 1", "Competitor 2", "Competitor 3", "Competitor 4", "Competitor 5"]
                for i, h in enumerate(headers, 1):
                    c = ws.cell(row=current_row, column=i, value=h); apply_style(c, font=f_tbl_hdr, fill=fill_tbl_hdr, alignment=align_center)
                current_row += 1
                
                for fld in ["ASIN", "Product Title"]:
                    c = ws.cell(row=current_row, column=1, value=fld); apply_style(c, font=f_side_lbl, fill=fill_side_lbl, alignment=align_left)
                    c = ws.cell(row=current_row, column=2, value=""); apply_style(c, fill=fill_highlight, alignment=align_center)
                    for i in range(3, 8):
                        c = ws.cell(row=current_row, column=i, value=""); apply_style(c, alignment=align_center)
                    current_row += 1
                
                current_row += 1
                headers[0] = "Comparison Metrics"
                for i, h in enumerate(headers, 1):
                    c = ws.cell(row=current_row, column=i, value=h); apply_style(c, font=f_tbl_hdr, fill=fill_tbl_hdr, alignment=align_center)
                current_row += 1
                
                for m in range(1, 6):
                    c = ws.cell(row=current_row, column=1, value=f"Metric {m} Name"); apply_style(c, font=f_side_lbl, fill=fill_side_lbl, alignment=align_left)
                    c = ws.cell(row=current_row, column=2, value=""); apply_style(c, fill=fill_highlight, alignment=align_center)
                    for i in range(3, 8):
                        c = ws.cell(row=current_row, column=i, value=""); apply_style(c, alignment=align_center)
                    current_row += 1
                    
            else:
                # Generic Module Layout
                current_row += 1
                has_repeat = any(f.get("repeat", 1) > 1 for f in mod["fields"])
                
                if has_repeat:
                    repeats = max([f.get("repeat", 1) for f in mod["fields"]])
                    headers = ["Fields"] + [f"Block {i}" for i in range(1, repeats + 1)]
                    for i, h in enumerate(headers, 1):
                        c = ws.cell(row=current_row, column=i, value=h); apply_style(c, font=f_tbl_hdr, fill=fill_tbl_hdr, alignment=align_center)
                    current_row += 1
                    
                    # We map fields by their logical group (ignoring the repeat count)
                    rendered_fields = set()
                    for fld in mod["fields"]:
                        f_name = fld["label"]
                        if fld.get("group") or fld.get("repeat", 1) > 1:
                            if f_name in rendered_fields:
                                continue
                            rendered_fields.add(f_name)
                            c = ws.cell(row=current_row, column=1, value=f_name); apply_style(c, font=f_side_lbl, fill=fill_side_lbl, alignment=align_left)
                            for i in range(1, fld.get("repeat", 1) + 1):
                                c = ws.cell(row=current_row, column=i+1, value=True if fld["type"] == "boolean" else ""); apply_style(c, fill=fill_config if fld["type"] == "boolean" else None, alignment=align_center)
                            current_row += 1
                        else:
                            c = ws.cell(row=current_row, column=1, value=f_name); apply_style(c, font=f_side_lbl, fill=fill_side_lbl, alignment=align_left)
                            c = ws.cell(row=current_row, column=2, value=True if fld["type"] == "boolean" else ""); apply_style(c, fill=fill_config if fld["type"] == "boolean" else None, alignment=align_center)
                            current_row += 1
                else:
                    for fld in mod["fields"]:
                        c = ws.cell(row=current_row, column=1, value=fld["label"]); apply_style(c, font=f_side_lbl, fill=fill_side_lbl, alignment=align_left)
                        c = ws.cell(row=current_row, column=2, value=True if fld["type"] == "boolean" else ""); apply_style(c, fill=fill_config if fld["type"] == "boolean" else None, alignment=align_center)
                        current_row += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    headers = {
        'Content-Disposition': 'attachment; filename="APlus_Template.xlsx"'
    }
    return Response(content=output.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)
